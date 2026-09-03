"""Composite Returns: upload-driven portfolio engine and Streamlit interface.
Run: streamlit run app.py
All new results are computed from source prices, never from saved Summary cells.
"""
from __future__ import annotations

import hashlib
import io
import math
import posixpath
import re
import streamlit as st
import zipfile
from dataclasses import dataclass
from datetime import date
from xml.etree import ElementTree as ET

import numpy as np
import pandas as pd

NS = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
PERIODS = [('1 Month', 1), ('3 Months', 3), ('6 Months', 6), ('1 Year', 12),
           ('3 Years', 36), ('5 Years', 60), ('10 Years', 120), ('Since Inception', None)]
FREQUENCIES = ['Monthly', 'Quarterly', 'Semi-Annually', 'Annually', 'Custom', 'Buy and Hold']
VOL_MODES = ['Match saved report: sample volatility', 'Match live formulas: population portfolio / sample indices']


class InputError(ValueError):
    pass


class ExcelSource:
    """Read OOXML cells without loading or executing VBA or Excel calculations."""
    def __init__(self, content: bytes):
        try:
            self.z = zipfile.ZipFile(io.BytesIO(content))
            if sum(x.file_size for x in self.z.infolist()) > 1_500_000_000:
                raise InputError('The expanded workbook is too large (limit 1.5 GB).')
            root = ET.fromstring(self.z.read('xl/workbook.xml'))
            rels = ET.fromstring(self.z.read('xl/_rels/workbook.xml.rels'))
            targets = {r.get('Id'): r.get('Target') for r in rels}
            self.sheets = {}
            self.states = {}
            for sheet in root.find('m:sheets', NS):
                target = targets[sheet.get('{' + REL + '}id')]
                self.sheets[sheet.get('name')] = (target.lstrip('/') if target.startswith('/')
                    else posixpath.normpath(posixpath.join('xl', target)))
                self.states[sheet.get('name')] = sheet.get('state', 'visible')
            props = root.find('m:workbookPr', NS)
            self.epoch = pd.Timestamp('1904-01-01' if props is not None and props.get('date1904') in ('1','true') else '1899-12-30')
            self.strings = []
            if 'xl/sharedStrings.xml' in self.z.namelist():
                self.strings = [''.join(t.text or '' for t in item.findall('.//m:t', NS))
                                for item in ET.fromstring(self.z.read('xl/sharedStrings.xml'))]
        except (zipfile.BadZipFile, KeyError, ET.ParseError) as exc:
            raise InputError('Upload an unencrypted .xlsm or .xlsx workbook in the Composite Returns layout.') from exc

    def cells(self, sheet, first_row=1, last_row=None):
        if sheet not in self.sheets:
            raise InputError(f'Missing required worksheet: {sheet}')
        result = {}
        with self.z.open(self.sheets[sheet]) as stream:
            for _, row in ET.iterparse(stream, events=('end',)):
                if row.tag != '{' + NS['m'] + '}row':
                    continue
                number = int(row.get('r'))
                if last_row is not None and number > last_row:
                    break
                if number >= first_row:
                    for cell in row:
                        v = cell.find('m:v', NS)
                        value = None if v is None else v.text
                        kind = cell.get('t')
                        if kind == 's' and value is not None:
                            value = self.strings[int(value)]
                        elif kind == 'inlineStr':
                            value = ''.join(t.text or '' for t in cell.findall('.//m:t', NS))
                        elif kind == 'b':
                            value = value == '1'
                        elif kind not in ('str', 'e', 'd') and value is not None:
                            try:
                                value = float(value)
                            except ValueError:
                                pass
                        if value is not None:
                            result[cell.get('r')] = value
                row.clear()
        return result

    def day(self, value):
        if isinstance(value, (int, float, np.number)) and np.isfinite(value):
            return (self.epoch + pd.Timedelta(days=float(value))).normalize()
        if value is None:
            raise InputError('A required input date is blank.')
        try:
            return pd.Timestamp(value).normalize()
        except (ValueError, TypeError) as exc:
            raise InputError(f'Invalid date: {value}') from exc


def number(value, label):
    try:
        val = float(value)
    except (ValueError, TypeError) as exc:
        raise InputError(f'{label} must contain a number, not {value!r}.') from exc
    if not np.isfinite(val):
        raise InputError(f'{label} must be finite.')
    return val


def header_columns(cells, row=1):
    return {k[:-len(str(row))]: str(v).strip() for k, v in cells.items()
            if re.fullmatch(r'[A-Z]+' + str(row), k) and v is not None}


def price_frame(cells, dates, headings):
    columns = {}
    for col, name in headings.items():
        if col == 'A':
            continue
        if name in columns:
            raise InputError(f'Duplicate source column: {name}')
        columns[name] = pd.to_numeric(pd.Series([cells.get(f'{col}{r}') for r in dates],
                                             index=list(dates.values())), errors='coerce')
    frame = pd.DataFrame(columns)
    if frame.index.has_duplicates:
        raise InputError('Source data contains duplicate dates. Resolve them before calculating.')
    return frame.sort_index()


def read_model(content):
    book = ExcelSource(content)
    for sheet in ['ETF Allocation', 'NAV', 'Index', 'Face Value', 'Holidays', 'Reblancing Dates']:
        if sheet not in book.sheets:
            raise InputError(f'Missing required worksheet: {sheet}')
    allocation = book.cells('ETF Allocation')
    weights = []
    for k, name in allocation.items():
        if re.fullmatch(r'A\d+', k) and int(k[1:]) >= 2 and isinstance(name, str):
            weight = allocation.get('B' + k[1:])
            if weight is not None:
                weights.append({'Index': name.strip(), 'Weight (%)': number(weight, name + ' weight') * 100})
    weights = pd.DataFrame(weights)
    if weights.empty or weights['Index'].duplicated().any():
        raise InputError('Allocation names must be present and unique in columns A:B.')
    nav_cells = book.cells('NAV')
    dates = {int(k[1:]): book.day(v) for k, v in nav_cells.items()
             if re.fullmatch(r'A\d+', k) and int(k[1:]) >= 3 and isinstance(v, (int, float))}
    if not dates:
        raise InputError('NAV column A must contain the source trading dates.')
    nav = price_frame(nav_cells, dates, header_columns(nav_cells))
    index_cells = book.cells('Index')
    # Row 2 is the original benchmark label; row 1 sometimes links to row 2.
    index_headers = header_columns(index_cells, 2)
    index = price_frame(index_cells, dates, index_headers)
    if nav.empty or index.empty:
        raise InputError('No usable source prices found.')
    for name, data in [('NAV', nav), ('Index', index)]:
        if (data.notna() & (data <= 0)).any().any():
            raise InputError(f'{name} contains zero or negative source prices.')
    face_cells = book.cells('Face Value')
    face = {v: number(face_cells.get('B'+k[1:]), f'Face value for {v}')
            for k,v in face_cells.items() if re.fullmatch(r'A\d+', k) and int(k[1:]) >= 2}
    # First-date metadata is a model assumption, not a saved portfolio calculation.
    declared = {'Index': {}, 'NAV': {}}
    for mode, sheet, fallback in [('Index', 'Adjusted Index Value', {}), ('NAV', 'Adjusted NAV', nav_cells)]:
        metadata = book.cells(sheet, last_row=2) if sheet in book.sheets else fallback
        for col, name in header_columns(metadata).items():
            val = metadata.get(col+'2')
            if col != 'A' and isinstance(val, (int, float)):
                declared[mode][name] = book.day(val)
    holiday = book.cells('Holidays')
    calendar = pd.DatetimeIndex(sorted({book.day(v) for k,v in holiday.items()
                  if re.fullmatch(r'F\d+', k) and isinstance(v, (int,float))}))
    if len(calendar) == 0:
        raise InputError('Holidays column F must contain the working-day calendar.')
    seeds = book.cells('Reblancing Dates', last_row=2)
    seed_dates = {freq: book.day(seeds.get(col+'2')) for freq,col in
                  [('Monthly','A'),('Quarterly','B'),('Semi-Annually','C'),('Annually','D'),('Custom','E')]}
    defaults = {
        'amount': number(allocation.get('F7'), 'Investment amount'),
        'start': book.day(allocation.get('F3')).date(),
        'end': book.day(allocation.get('F10')).date(),
        'mode': str(allocation.get('F9', 'Index')),
        'frequency': str(allocation.get('F5', 'Annually')),
        'custom_step': int(number(allocation.get('G5', 12), 'Custom interval')),
        'custom_unit': str(allocation.get('G6', 'Months')),
    }
    return {'raw': {'Index':index,'NAV':nav}, 'face':face, 'declared':declared,
            'calendar':calendar, 'seeds':seed_dates, 'allocation':weights,
            'defaults':defaults, 'sheets':list(book.sheets), 'states':book.states}


def previous_working(calendar, value):
    value = pd.Timestamp(value).normalize()
    loc = calendar.searchsorted(value, side='right') - 1
    if loc < 0:
        return None
    return calendar[loc]


def next_working(calendar, value):
    loc = calendar.searchsorted(pd.Timestamp(value).normalize(), side='left')
    if loc >= len(calendar):
        raise InputError('Start date is beyond the working-day calendar.')
    return calendar[loc]


def adjusted_prices(model, mode, names):
    if mode not in ('Index','NAV'):
        raise InputError('Series based on must be Index or NAV.')
    raw = model['raw'][mode]
    missing = [n for n in names if n not in raw]
    if missing:
        raise InputError(f'{mode} data is unavailable for: ' + ', '.join(missing) + '. Change the allocation or series choice.')
    output = pd.DataFrame(index=raw.index)
    starts = {}
    fills = {}
    for name in names:
        source = raw[name]
        first = source.first_valid_index()
        if first is None:
            raise InputError(f'No numeric {mode} data for {name}.')
        declared = model['declared'][mode].get(name, first)
        if declared not in source.index or pd.isna(source.at[declared]):
            raise InputError(f'{name}: declared first date {declared.date()} has no source price.')
        starts[name] = max(first, declared)
        adjusted = source.ffill()
        adjusted.loc[adjusted.index < starts[name]] = np.nan
        fills[name] = int((source.isna() & adjusted.notna()).sum())
        if mode == 'Index':
            base = model['face'].get(name)
            if base is None or base <= 0:
                raise InputError(f'Positive Face Value missing for {name}.')
            adjusted = adjusted / source.at[declared] * base
        output[name] = adjusted
    return output, starts, fills


def rebalancing_dates(model, frequency, end, step=12, unit='Months'):
    if frequency == 'Buy and Hold':
        return pd.DatetimeIndex([])
    if frequency not in FREQUENCIES:
        raise InputError('Unsupported rebalancing frequency.')
    if frequency == 'Custom' and (step < 1 or unit not in ('Days','Months')):
        raise InputError('Custom interval must be a positive number of Days or Months.')
    months = {'Monthly':1,'Quarterly':3,'Semi-Annually':6,'Annually':12}
    cursor = model['seeds'][frequency]
    dates = [cursor]
    end = pd.Timestamp(end)
    for _ in range(50000):
        if cursor >= end:
            break
        if frequency == 'Custom' and unit == 'Days':
            # Advance the nominal date; do not get trapped on a weekend after rollback.
            nominal = cursor + pd.Timedelta(days=step)
            mapped = previous_working(model['calendar'], nominal)
            cursor = nominal
        else:
            # EOMONTH(d,n) always targets n calendar months after d's month.
            nominal = (cursor + pd.DateOffset(months=months.get(frequency,step))) + pd.offsets.MonthEnd(0)
            mapped = previous_working(model['calendar'], nominal)
            if mapped is None or mapped <= cursor:
                break
            cursor = mapped
        if mapped is not None and mapped <= end:
            dates.append(mapped)
    else:
        raise InputError('Custom schedule exceeds 50,000 periods.')
    return pd.DatetimeIndex(sorted(set(dates)))


@dataclass
class Simulation:
    daily: pd.DataFrame
    units: pd.DataFrame
    market_values: pd.DataFrame
    weights: pd.DataFrame
    trades: pd.DataFrame


def simulate(prices, weights, amount, schedule):
    if prices.empty or prices.isna().any().any() or (prices <= 0).any().any():
        raise InputError('Selected period has missing or invalid adjusted prices.')
    names = list(weights.index)
    p = prices[names].to_numpy(dtype=float)
    dates = prices.index
    target = weights.to_numpy(dtype=float)
    held = np.zeros(len(names)); cash = float(amount); total = float(amount)
    daily=[]; unit_rows=[]; mv_rows=[]; weight_rows=[]; trades=[]
    schedule = set(schedule)
    for i, day in enumerate(dates):
        if i > 0 and (i == 1 or dates[i-1] in schedule):
            old = held.copy()
            held = np.floor(total * target / p[i-1])
            cash = total - float(held @ p[i-1])
            for j,name in enumerate(names):
                trades.append({'Pricing date':dates[i-1], 'Effective date':day,
                    'Event':'Initial investment' if i==1 else 'Rebalance', 'Index':name,
                    'Portfolio before (₹)':total, 'Target weight (%)':target[j]*100,
                    'Price used':p[i-1,j], 'Old units':old[j], 'New units':held[j],
                    'Unit change':held[j]-old[j], 'Trade value (₹)':(held[j]-old[j])*p[i-1,j],
                    'Cash after (₹)':cash})
        mv = held*p[i]
        total = float(amount) if i==0 else float(mv.sum()+cash)
        daily.append({'Date':day,'Portfolio value (₹)':total,'Cash (₹)':cash})
        unit_rows.append(held.copy());mv_rows.append(mv.copy())
        weight_rows.append(np.append(mv/total, cash/total))
    daily = pd.DataFrame(daily).set_index('Date')
    daily['Daily return'] = daily['Portfolio value (₹)'].pct_change(fill_method=None)
    return Simulation(daily, pd.DataFrame(unit_rows,index=dates,columns=names),
        pd.DataFrame(mv_rows,index=dates,columns=names),
        pd.DataFrame(weight_rows,index=dates,columns=names+['Cash']),pd.DataFrame(trades))


def return_rate(start_value, end_value, days):
    if days <= 0 or start_value <= 0:
        return np.nan
    ratio = end_value/start_value
    return ratio-1 if days<=365 else ratio**(365/days)-1


def summary_tables(model, selected, buyhold, prices, amount, requested_end, vol_mode):
    start, end = prices.index[0],prices.index[-1]
    anchors={}
    requested_end=pd.Timestamp(requested_end)
    for label,months in PERIODS:
        if months is None:
            anchor=start
        else:
            nominal=requested_end-pd.DateOffset(months=months)
            if requested_end.is_month_end:
                nominal=nominal+pd.offsets.MonthEnd(0)
            anchor=previous_working(model['calendar'],nominal)
        anchors[label]=anchor
    rows=[]; kinds=[]
    def block(label, series, daily_returns, portfolio, ddof):
        metrics={key:{} for key in (['Return','Value of investment (₹)','Annualised volatility','Return / Risk']
                                    if portfolio else ['Return','Annualised volatility','Return / Risk'])}
        for period,_ in PERIODS:
            anchor=anchors[period]
            valid=anchor is not None and start<=anchor<end and anchor in series.index
            ret=vol=value=risk=np.nan
            if valid:
                ret=return_rate(float(series.at[anchor]),float(series.at[end]),(end-anchor).days)
                observations=daily_returns.loc[(daily_returns.index>anchor)&(daily_returns.index<=end)].dropna()
                vol=float(observations.std(ddof=ddof)*np.sqrt(250)) if len(observations)>ddof else np.nan
                risk=ret/vol if pd.notna(vol) and vol>0 else np.nan
                value=float(series.at[end]/series.at[anchor]*amount)
            metrics['Return'][period]=ret
            metrics['Annualised volatility'][period]=vol
            metrics['Return / Risk'][period]=risk
            if portfolio:
                metrics['Value of investment (₹)'][period]=value
        for metric,values in metrics.items():
            rows.append({'Particulars':label+' · '+metric,**values});kinds.append(metric)
    portfolio_ddof=1 if vol_mode==VOL_MODES[0] else 0
    block('Portfolio — selected frequency',selected.daily['Portfolio value (₹)'],selected.daily['Daily return'],True,portfolio_ddof)
    block('Portfolio — buy and hold',buyhold.daily['Portfolio value (₹)'],buyhold.daily['Daily return'],True,portfolio_ddof)
    constituent=prices.pct_change(fill_method=None)
    for name in prices.columns:
        block(name,prices[name],constituent[name],False,1)
    table=pd.DataFrame(rows).set_index('Particulars')
    display=table.copy().astype(object)
    for i,kind in enumerate(kinds):
        for col in display.columns:
            value=table.iloc[i][col]
            display.iloc[i,display.columns.get_loc(col)]=('—' if pd.isna(value) else
                f'₹{value:,.2f}' if kind=='Value of investment (₹)' else
                f'{value:.2f}' if kind=='Return / Risk' else f'{value:.2%}')
    windows=pd.DataFrame([{'Period':label,'Start date':anchor,'End date':end,
        'Calendar days':(end-anchor).days if anchor is not None else None,
        'Available':anchor is not None and start<=anchor<end and anchor in prices.index}
        for label,anchor in anchors.items()])
    return table,display,windows,constituent


def calculate(model, config, allocation, vol_mode=VOL_MODES[0]):
    amount=number(config['amount'],'Investment amount')
    if amount <= 0:
        raise InputError('Investment amount must be greater than zero.')
    w=allocation.copy()
    w['Weight (%)']=pd.to_numeric(w['Weight (%)'],errors='coerce')
    if w['Weight (%)'].isna().any() or not np.isfinite(w['Weight (%)']).all() or (w['Weight (%)']<0).any():
        raise InputError('Weights must be finite non-negative numbers.')
    if not math.isclose(float(w['Weight (%)'].sum()),100.0,abs_tol=1e-8,rel_tol=0):
        raise InputError(f'Weights total {w["Weight (%)"].sum():.6f}%. They must total 100%.')
    weights=w.loc[w['Weight (%)']>0].set_index('Index')['Weight (%)']/100
    adjusted,starts,fills=adjusted_prices(model,config['mode'],list(weights.index))
    requested_start=pd.Timestamp(config['start']);requested_end=pd.Timestamp(config['end'])
    if requested_start>=requested_end:
        raise InputError('End date must be later than investment date.')
    if requested_end>adjusted.index.max():
        raise InputError(f'The uploaded data ends on {adjusted.index.max():%d-%b-%Y}. Choose an end date on or before that date.')
    effective_start=next_working(model['calendar'],max(requested_start,max(starts.values())))
    effective_end=previous_working(model['calendar'],requested_end)
    if effective_end is None or effective_start>=effective_end:
        raise InputError('The selected investments do not have a common usable period between these dates.')
    if effective_start not in adjusted.index or effective_end not in adjusted.index:
        raise InputError('The working-day calendar does not align with the source price dates.')
    prices=adjusted.loc[effective_start:effective_end]
    expected=model['calendar'][(model['calendar']>=effective_start)&(model['calendar']<=effective_end)]
    if not expected.equals(prices.index):
        raise InputError('Source dates and the working-day calendar differ inside the selected period.')
    schedule=rebalancing_dates(model,config['frequency'],effective_end,config['custom_step'],config['custom_unit'])
    selected=simulate(prices,weights,amount,schedule)
    buyhold=simulate(prices,weights,amount,[])
    summary,display,windows,constituent=summary_tables(model,selected,buyhold,prices,amount,requested_end,vol_mode)
    metadata=pd.DataFrame([{'Index':n,'Target weight (%)':weights[n]*100,
        'Initial target allocation (₹)':weights[n]*amount,'First usable date':starts[n],
        'Forward-filled observations (all history)':fills[n]} for n in weights.index])
    reconciliation=selected.market_values.sum(axis=1)+selected.daily['Cash (₹)']-selected.daily['Portfolio value (₹)']
    checks=pd.DataFrame([
        {'Check':'Weights total 100%','Result':f'{weights.sum()*100:.8f}%','Status':'PASS'},
        {'Check':'Source prices available throughout period','Result':f'{len(prices):,} trading dates','Status':'PASS'},
        {'Check':'Holdings plus cash equals portfolio','Result':f'Max difference ₹{reconciliation.abs().max():.10f}','Status':'PASS' if reconciliation.abs().max()<1e-6 else 'FAIL'},
        {'Check':'Whole units and non-negative cash','Result':f'Min cash ₹{selected.daily["Cash (₹)"].min():.2f}','Status':'PASS' if selected.daily['Cash (₹)'].min()>=-1e-6 else 'FAIL'},
    ])
    return dict(selected=selected,buyhold=buyhold,summary=summary,display=display,windows=windows,
                prices=prices,adjusted=adjusted,constituent=constituent,metadata=metadata,
                schedule=schedule,checks=checks,start=effective_start,end=effective_end,
                config=config,vol_mode=vol_mode)


EXPLANATION = r"""
### 1. Source data and inputs
The uploaded **ETF Allocation** sheet supplies weights (A:B), investment date (F3),
frequency (F5), amount (F7), Index/NAV choice (F9), end date (F10), and custom interval
(G5:G6). The **Index** and **NAV** sheets supply historical observations. **Face Value**
supplies index bases; **Holidays F:F** supplies the working-day calendar. Declared first
usable dates come from the first-date metadata in the adjusted sheets.
Saved Summary results, Units, Market Value and Series are never used as calculation inputs.
Macros are never executed. Every displayed new result is calculated in Python.

### 2. Prepare prices
NAV mode uses the source ETF NAV. Index mode uses:

`Adjusted index price[t] = Raw index[t] / Raw index[first declared date] × Face Value`

After the first usable observation, a missing source value is carried forward from the
previous observation, matching the workbook. Missing leading history is not backfilled.
Changing the index base can affect whole-unit rounding and residual cash, so the workbook's
Face Value assumptions are retained. The source sheet's mapped ETF series is used in NAV mode.

### 3. Resolve dates
`Effective start = next working day on/after max(requested start, selected first usable dates)`

`Effective end = previous working day on/before requested end`

Dates after the uploaded source history are rejected. The app does not invent updated prices.

### 4. Build the rebalancing calendar
Monthly, quarterly, semi-annual and annual dates use the corresponding seed date from
**Reblancing Dates**, then add 1, 3, 6 or 12 months, take month-end and roll back to the
previous working day. Annual rebalancing in the supplied workbook is in December.
Buy and hold has no subsequent rebalances. Custom Months uses the workbook's seed and
month-end rule. Custom Days advances nominal dates by the chosen interval, rolls each back
to a working day and removes duplicates. This deliberately prevents the Excel custom-day
formula from getting stuck on a weekend/holiday when its adjusted date does not advance.

### 5. Allocate, round units and retain cash
For each selected constituent at inception and every rebalance:

`New units[i] = floor(portfolio value at pricing date × target weight[i] / price[i])`

`Cash = portfolio value at pricing date − sum(new units[i] × price[i])`

The initial date holds the entered portfolio value. Holdings priced on that date first
appear on the next trading date. Similarly, a rebalance priced at month/year-end takes effect
on the next trading date. This reproduces the original workbook's timing. Units remain
constant between rebalances. Cash remains uninvested and earns no interest.

### 6. Value the portfolio and calculate daily returns
`Holding market value[i,t] = units[i,t] × adjusted price[i,t]`

`Portfolio value[t] = sum(holding market values[t]) + cash[t]`

`Actual weight[i,t] = holding market value[i,t] / portfolio value[t]`

`Cash weight[t] = cash[t] / portfolio value[t]`

`Daily portfolio return[t] = portfolio value[t] / portfolio value[t−1] − 1`

`Daily constituent return[i,t] = price[i,t] / price[i,t−1] − 1`

The **Daily calculations** tab exposes prices, units, market values, cash, actual weights,
returns and the full initial-investment/rebalancing ledger for both strategies.

### 7. Choose summary windows
The report contains 1, 3 and 6 months; 1, 3, 5 and 10 years; and Since Inception.
For each lookback, subtract its calendar months from the requested end date. If that end
is month-end, preserve month-end when shifting. Roll the anchor back to the previous working
day. Since Inception starts at the effective investment date. Periods before inception are
unavailable and display **—**. They do not use a shortened period under the original label.
These are returns along the existing portfolio history, not fresh target-weight portfolios
launched independently for each lookback.

### 8. Returns and investment values
`Absolute return = ending value / starting value − 1` for durations up to 365 days.

`CAGR = (ending value / starting value) ** (365 / calendar days) − 1` beyond 365 days.

`Value of entered investment over window = entered amount × ending value / window starting value`

The same date and return rules apply to individual indices using their adjusted prices.
The app fixes the old constituent Since Inception formula's unconditional annualisation:
if inception is within a year, the constituent also uses absolute return. All constituents
follow the same day-count rule, including lookbacks that cross leap years.

### 9. Annualised volatility
`Annualised volatility = standard deviation(daily returns in window) × sqrt(250)`

Only daily returns strictly after the period start and through the end date are included.
**Match saved report** uses sample standard deviation (`ddof=1`, Excel STDEV) for both
portfolio strategies and constituents; this matches the uploaded saved report.
**Match live formulas** uses population standard deviation (`ddof=0`, Excel STDEVP) for
portfolios and sample standard deviation for constituents, matching the current helper
formulas. The original file mixes these conventions. There are 250 assumed trading days.
Insufficient observations produce **—**, rather than a misleading zero.

### 10. Return / Risk
`Return / Risk = displayed period return / annualised volatility`

This reproduces the report's ratio; it is not a Sharpe ratio, and no risk-free return is
subtracted. For short windows, absolute returns are divided by annualised volatility, as
in the workbook. Zero volatility produces **—**.

### 11. Buy and hold comparison
A second independent simulation uses the same source prices, effective dates, amount,
initial weights and whole-unit rounding, with no subsequent rebalancing. Both strategies
are calculated together; the app does not change an Excel cell or reuse stale macro output.

### 12. Model boundaries and validation
This is a single initial investment model: no SIPs, withdrawals, brokerage, taxes, slippage
or separate fee deductions are added. NAV inputs already reflect whatever is embedded in
the supplied NAV. TRI observations are used as supplied, without adding dividends again.
The workbook's ₹1 lakh amount message is advisory; this app accepts any positive amount and
shows the same advisory below ₹1 lakh. Weights must be non-negative and total 100%.
Unavailable NAV constituents are reported explicitly instead of silently zeroing weights.
"""

CSS = """
<style>
.stApp {background:#08090b;color:#f3eee2;}
[data-testid="stSidebar"] {background:#111216;border-right:1px solid #544527;}
h1,h2,h3 {color:#e3bd63 !important;} .stCaption {color:#c4bba7;}
[data-testid="stMetric"] {background:#151619;border:1px solid #58482b;border-radius:12px;padding:18px;}
[data-testid="stMetricValue"] {color:#f0cf83;}
.stButton>button, .stFormSubmitButton>button {background:#d6b05d;color:#101113;border:0;font-weight:700;}
.stButton>button:hover,.stFormSubmitButton>button:hover {background:#eed28e;color:#08090b;}
[data-baseweb="tab-list"] {gap:20px;border-bottom:1px solid #544527;}
button[data-baseweb="tab"] {color:#bcb5a6;font-size:16px;}
button[data-baseweb="tab"][aria-selected="true"] {color:#f0cf83;}
[data-baseweb="tab-highlight"] {background:#e3bd63;}
[data-testid="stFileUploader"] {border:1px dashed #a98a44;border-radius:12px;padding:12px;}
.hero {padding:10px 0 22px;border-bottom:1px solid #544527;margin-bottom:22px;}
.kicker {letter-spacing:3px;color:#bd9b50;font-size:12px;font-weight:700;}
</style>
"""


def show_table(st, frame, percent=False):
    styled=frame.style.set_properties(**{'background-color':'#111216','color':'#f2eada','border-color':'#403722'})
    if percent:
        styled=styled.format('{:.2%}',na_rep='—')
    st.dataframe(styled,width='stretch')


def original_page(content, sheet, start_row, count, formulas=False):
    # openpyxl resolves shared formulas and date formatting for the original-data viewer.
    from openpyxl import load_workbook
    wb=load_workbook(io.BytesIO(content),read_only=True,data_only=not formulas,keep_links=False)
    try:
        ws=wb[sheet]
        maxrow,maxcol=ws.max_row,ws.max_column
        rows=list(ws.iter_rows(min_row=start_row,max_row=min(start_row+count-1,maxrow),
                              max_col=maxcol,values_only=True))
        from openpyxl.utils import get_column_letter
        frame=pd.DataFrame(rows,index=range(start_row,start_row+len(rows)),
                           columns=[get_column_letter(i) for i in range(1,maxcol+1)])
        # Preserve mixed Excel values faithfully in a display grid without Arrow type errors.
        frame=frame.map(lambda v: '' if v is None else str(v))
        frame.index.name='Excel row'
        return frame,maxrow,maxcol
    finally:
        wb.close()


def format_summary_table(summary, amount, view):
    """Format calculated results without changing their values."""
    formatted = summary.copy().astype(object)
    labels = []

    if amount >= 100000:
        investment_label = (
            f"Value of Rs. {amount / 100000:g} lakh Invested"
        )
    else:
        investment_label = (
            f"Value of Rs. {amount:,.0f} Invested"
        )

    for row_number, label in enumerate(summary.index):
        entity, metric = str(label).rsplit(" · ", 1)

        if entity == "Portfolio — selected frequency":
            suffix = "after Periodic Rebalancing"
        elif entity == "Portfolio — buy and hold":
            suffix = "Buy and Hold - No Rebalancing"
        else:
            suffix = None

        if suffix is not None:
            row_name = {
                "Return": (
                    f"Portfolio Returns ({suffix})"
                ),
                "Value of investment (₹)": (
                    f"{investment_label} ({suffix})"
                ),
                "Annualised volatility": (
                    f"Annualized Volatility (%) ({suffix})"
                ),
                "Return / Risk": (
                    f"Return/ Risk ({suffix})"
                ),
            }[metric]
        else:
            row_name = {
                "Return": (
                    f"{entity} Returns"
                ),
                "Annualised volatility": (
                    f"{entity} Annualized Volatility (%)"
                ),
                "Return / Risk": (
                    f"{entity} Return/ Risk"
                ),
            }[metric]

        if view == "Decimal values":
            row_name = row_name.replace(
                "Volatility (%)",
                "Volatility (decimal)",
            )

        labels.append(row_name)

        for column_number in range(len(summary.columns)):
            value = summary.iloc[
                row_number,
                column_number,
            ]

            if pd.isna(value):
                text = "-"

            elif metric == "Value of investment (₹)":
                text = f"{value:,.0f}"

            elif metric == "Return / Risk":
                text = f"{value:.2f}"

            elif view == "Percentages":
                text = f"{value:.2%}"

            else:
                text = (
                    f"{value:.8f}"
                    .rstrip("0")
                    .rstrip(".")
                )

            formatted.iloc[
                row_number,
                column_number,
            ] = text

    formatted.index = pd.Index(
        labels,
        name="Particulars",
    )

    return formatted



def main():
    import streamlit as st
    st.set_page_config(page_title='Composite Returns | Portfolio Lab',page_icon='◈',layout='wide')
    st.markdown(CSS,unsafe_allow_html=True)
    st.markdown('<div class="hero"><div class="kicker">PORTFOLIO LAB</div><h1>Composite Returns</h1><p>Historical portfolio performance, calculated from your workbook.</p></div>',unsafe_allow_html=True)
    uploaded=st.file_uploader('Upload the Composite Returns workbook',type=['xlsm','xlsx'])
    if uploaded is None:
        st.info('Upload your Excel workbook to unlock the portfolio inputs, summary, original data and calculation explanations.')
        st.stop()
    content=uploaded.getvalue()
    fingerprint=hashlib.sha256(content).hexdigest()
    @st.cache_data(show_spinner=False,max_entries=2)
    def cached_model(data):
        return read_model(data)
    @st.cache_data(show_spinner=False,max_entries=4)
    def cached_page(data,sheet,row,count,formulas):
        return original_page(data,sheet,row,count,formulas)
    try:
        with st.spinner('Reading source prices and workbook inputs…'):
            model=cached_model(content)
    except (InputError,ValueError,KeyError,zipfile.BadZipFile) as exc:
        st.error(str(exc));st.stop()
    if st.session_state.get('upload_id')!=fingerprint:
        for key in list(st.session_state):
            if key.startswith('input_'):
                del st.session_state[key]
        st.session_state['upload_id']=fingerprint
    defaults=model['defaults']
    with st.sidebar:
        st.header('Portfolio inputs')
        st.caption('Defaults are read from the uploaded workbook. Submit changes to recalculate all results.')
        with st.form('portfolio_inputs'):
            amount=st.number_input('Investment amount (₹)',min_value=1.0,value=max(1.0,defaults['amount']),step=10000.0,key='input_amount')
            start=st.date_input('Investment date',value=defaults['start'],min_value=date(1900,1,1),max_value=date(2100,12,31),key='input_start')
            end=st.date_input('As on date',value=defaults['end'],min_value=date(1900,1,1),max_value=date(2100,12,31),key='input_end')
            mode=st.selectbox('Series based on',['Index','NAV'],index=0 if defaults['mode']=='Index' else 1,key='input_mode')
            frequency=st.selectbox('Rebalancing frequency',FREQUENCIES,index=FREQUENCIES.index(defaults['frequency']) if defaults['frequency'] in FREQUENCIES else 3,key='input_frequency')
            st.caption('Custom interval below is used only when frequency is Custom.')
            step=st.number_input('Custom interval',min_value=1,max_value=10000,value=max(1,min(10000,defaults['custom_step'])),key='input_step')
            unit = st.selectbox(
                "Custom interval unit",
                ["Months", "Days"],
                index=(
                    1
                    if defaults["custom_unit"] == "Days"
                    else 0
                ),
                key="input_unit",
            )

            vol_mode = st.selectbox(
                "Volatility calculation",
                VOL_MODES,
                key="input_vol",
            )

            st.caption(
                f"Available indices: {len(model['allocation'])}. "
                "Enter weights below; the total must equal 100%."
            )

            allocation = st.data_editor(
                model["allocation"],
                hide_index=True,
                disabled=["Index"],
                column_config={
                    "Index": st.column_config.TextColumn(
                        "Name of Index",
                        width="large",
                    ),
                    "Weight (%)": st.column_config.NumberColumn(
                        "Weight in Portfolio (%)",
                        min_value=0.0,
                        max_value=100.0,
                        step=1.0,
                        format="%.2f",
                    ),
                },
                width="stretch",
                height=(len(model["allocation"]) + 1) * 35 + 3,
                key="input_allocation",
            )

            st.form_submit_button(
                "Calculate portfolio",
                width="stretch",
            )

    config = dict(
        amount=amount,
        start=start,
        end=end,
        mode=mode,
        frequency=frequency,
        custom_step=int(step),
        custom_unit=unit,
    )

    summary_tab, data_tab, explanation_tab, daily_tab = st.tabs(
        [
            "Summary",
            "Original Data",
            "Calculation Explanation",
            "Daily calculations",
        ]
    )

    result = None
    try:
        with st.spinner('Calculating both portfolio strategies…'):
            result=calculate(model,config,allocation,vol_mode)
    except (InputError,ValueError,KeyError) as exc:
        with summary_tab:
            st.error(str(exc))
            st.info('Correct the inputs and click Calculate portfolio. Original Data remains available.')
    if result is not None:
        with summary_tab:
            s=result['selected'];b=result['buyhold'];table=result['summary']
            st.caption(f'{frequency} · {mode} series · {result["start"]:%d-%b-%Y} to {result["end"]:%d-%b-%Y} · {len(s.daily):,} trading dates')
            if result['start']!=pd.Timestamp(start):
                st.info(f'Effective investment date: {result["start"]:%d-%b-%Y}, after working-day and source-history adjustments.')
            if result['end']!=pd.Timestamp(end):
                st.info(f'Effective end date: {result["end"]:%d-%b-%Y}, the preceding working day.')
            if amount<100000:
                st.warning('The workbook recommends an investment of at least ₹1 lakh. Smaller amounts can create more rounding cash.')
            metrics=st.columns(4)
            metrics[0].metric('Initial investment',f'₹{amount:,.0f}')
            metrics[1].metric('Portfolio value',f'₹{s.daily.iloc[-1]["Portfolio value (₹)"]:,.2f}')
            value=table.iloc[0]['Since Inception']
            metrics[2].metric('CAGR' if (result['end']-result['start']).days>365 else 'Absolute return',f'{value:.2%}')
            difference=s.daily.iloc[-1]['Portfolio value (₹)']-b.daily.iloc[-1]['Portfolio value (₹)']
            metrics[3].metric('Value versus buy and hold',f'₹{difference:+,.2f}')
            st.subheader("Allocation")
            show_table(st, result["metadata"])

            st.subheader(
                f"Performance as on "
                f"{result['end']:%d-%b-%Y}"
            )

            summary_view = st.radio(
                "Display returns and volatility as",
                ["Percentages", "Decimal values"],
                horizontal=True,
                key="summary_number_view",
            )

            st.caption(
                "Investment values are in rupees "
                "and rounded to whole numbers. "
                "Return/Risk is shown to two decimal places. "
                "Unavailable periods display -."
            )

            displayed_summary = format_summary_table(
                result["summary"],
                amount,
                summary_view,
            )

            show_table(st, displayed_summary)

            st.download_button(
                "Download summary in selected format (CSV)",
                displayed_summary.to_csv().encode(
                    "utf-8-sig"
                ),
                "formatted_summary.csv",
                "text/csv",
            )

            with st.expander("Portfolio value over time"):
                trend = pd.DataFrame(
                    {
                        "Selected frequency": (
                            result["selected"].daily[
                                "Portfolio value (₹)"
                            ]
                        ),
                        "Buy and hold": (
                            result["buyhold"].daily[
                                "Portfolio value (₹)"
                            ]
                        ),
                    }
                )

                st.line_chart(
                    trend,
                    color=["#d7b563", "#93979e"],
                )

            with st.expander("Period dates and validation"):
                show_table(st, result["windows"])
                show_table(st, result["checks"])

            st.caption(
                "All results above are calculated from "
                "source prices in this upload. "
                "The workbook's saved Summary is available "
                "in Original Data for comparison."
            )


    with data_tab:
        st.subheader('Original workbook data')
        st.caption('Browse all visible and hidden sheets. Values are the original saved Excel values; formula view shows the original formulas. These are not newly calculated results.')
        sheet=st.selectbox('Worksheet',model['sheets'],key='original_sheet')
        st.caption(f'Original worksheet visibility: {model["states"][sheet]}')
        cols=st.columns(3)
        row=int(cols[0].number_input('Starting Excel row',min_value=1,value=1,step=200,key='original_row'))
        count=cols[1].selectbox('Rows per page',[50,100,200,500],index=2)
        formulas=cols[2].checkbox('Show original formulas')
        with st.spinner('Reading original worksheet…'):
            frame,maxrow,maxcol=cached_page(content,sheet,row,count,formulas)
        st.caption(f'{maxrow:,} worksheet rows · {maxcol} columns. Original Excel row numbers and column letters are retained.')
        show_table(st,frame)
        st.download_button('Download this original-data page (CSV)',frame.to_csv().encode(),
                           'original_data_page.csv','text/csv')
        with st.expander('Download complete source price tables'):
            for source in ['Index','NAV']:
                st.download_button(f'Download complete {source} source (CSV)',model['raw'][source].to_csv().encode(),f'original_{source.lower()}_prices.csv','text/csv')
    with explanation_tab:
        st.markdown(EXPLANATION)
        if result is not None:
            st.subheader('Your current calculation')
            endvalue=result['selected'].daily.iloc[-1]['Portfolio value (₹)']
            days=(result['end']-result['start']).days
            equation=f'({endvalue:.8f} / {amount:.8f}) ** (365 / {days}) - 1' if days>365 else f'{endvalue:.8f} / {amount:.8f} - 1'
            st.code(equation,language='python')
            st.write(f'Result: **{result["summary"].iloc[0]["Since Inception"]:.6%}**')
            st.write(f'Volatility convention: **{vol_mode}**. Annualisation: **250 trading days**.')
            show_table(st,result['windows'])
    with daily_tab:
        if result is None:
            st.info('A valid allocation and date range are required for daily calculations.')
        else:
            strategy=st.radio('Strategy',['Selected frequency','Buy and hold'],horizontal=True)
            sim=result['selected'] if strategy=='Selected frequency' else result['buyhold']
            tables={'Portfolio value, cash and daily return':sim.daily,
                'Adjusted prices':result['prices'],'Units held':sim.units,
                'Holding market values (₹)':sim.market_values,
                'Holdings valued at previous-day prices (₹)':(sim.units*result['prices'].shift(1)).fillna(0),
                'Actual weights (%)':sim.weights*100,
                'Scheduled rebalancing dates':pd.DataFrame({'Pricing date':result['schedule'][(result['schedule']>=result['start']) & (result['schedule']<=result['end'])]}) if strategy=='Selected frequency' else pd.DataFrame({'Pricing date':[]}),
                'Constituent daily returns (%)':result['constituent']*100,
                'Initial investment and rebalancing ledger':sim.trades}
            dataset=st.selectbox('Calculation table',list(tables))
            frame=tables[dataset]
            page=int(st.number_input('Calculation page',min_value=1,max_value=max(1,math.ceil(len(frame)/200)),value=1))
            st.caption(f'{len(frame):,} total rows · 200 rows per page. Download includes all rows.')
            show_table(st,frame.iloc[(page-1)*200:page*200])
            st.download_button('Download complete calculation table (CSV)',frame.to_csv().encode(),
                               'portfolio_calculations.csv','text/csv')


if __name__=='__main__':
    main()
